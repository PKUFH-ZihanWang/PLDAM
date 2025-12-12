import os
import uuid
import nrrd
import pandas as pd
import numpy as np
from radiomics import featureextractor
from openpyxl import load_workbook
from concurrent.futures import ProcessPoolExecutor

# ---------------------------
# Utility Functions
# ---------------------------

def find_3dSlice_files(root_directory):
    """Find subdirectories containing '(P' and list all files."""
    dir_files = {}
    for dirpath, _, filenames in os.walk(root_directory):
        if '(P' in dirpath:
            full_paths = [os.path.join(dirpath, f) for f in filenames]
            dir_files[dirpath] = full_paths
    return dir_files


def append_data_to_excel(file_path, df, start_row, is_first_write=False):
    """Append DataFrame to Excel file."""
    df = df.astype(str)
    if not os.path.exists(file_path):
        pd.DataFrame(columns=df.columns).to_excel(file_path, index=False)
    workbook = load_workbook(file_path)
    sheet = workbook.active

    if is_first_write:
        for col_index, column in enumerate(df.columns, start=1):
            sheet.cell(row=1, column=col_index, value=column)
        start_row += 1

    for r in range(len(df)):
        for c in range(len(df.columns)):
            sheet.cell(row=start_row + r, column=c + 1, value=df.iat[r, c])

    workbook.save(file_path)


def extract_radiomics_features(mask_path, image_path, first_row, modality, hospital, writer_path, is_fact, start_row):
    """Extract radiomics features and append to Excel."""
    try:
        image_data, image_header = nrrd.read(image_path)
        mask_data, mask_header = nrrd.read(mask_path)

        # Unique temp file names
        temp_id = str(uuid.uuid4())
        temp_image = f"temp_image_{temp_id}.nrrd"
        temp_mask = f"temp_mask_{temp_id}.nrrd"

        nrrd.write(temp_image, np.clip(image_data, a_min=0, a_max=None), image_header)
        nrrd.write(temp_mask, mask_data.astype(int), image_header)

        settings = {
            'minimumROIDimensions': 3,
            'normalize': False,
            'resampledPixelSpacing': None,
            'distances': [1, 2, 3, 4, 5],
            'interpolator': 'sitkBSpline',
            'force2D': False,
        }
        extractor = featureextractor.RadiomicsFeatureExtractor(**settings)
        result = extractor.execute(temp_image, temp_mask)

        result_df = pd.DataFrame([result])
        result_df['ImagePath'] = image_path
        result_df['MaskPath'] = mask_path
        result_df['hospital'] = hospital
        result_df['modality'] = modality
        result_df['Pelvic'] = is_fact

        append_data_to_excel(writer_path, result_df, start_row=start_row, is_first_write=first_row)

        os.remove(temp_image)
        os.remove(temp_mask)

    except Exception as e:
        result_df = pd.DataFrame([{
            'ImagePath': image_path,
            'MaskPath': mask_path,
            'hospital': hospital,
            'modality': modality,
            'Pelvic': is_fact,
            'Error': str(e)
        }])
        append_data_to_excel(writer_path, result_df, start_row=start_row, is_first_write=first_row)
        print(f"[Error] Feature extraction failed: {e}")


def writer_map(hospital, modality, base_output_dir="/mnt/external_drive/RadiomicsResults"):
    """Return Excel output paths for each ROI."""
    output_dir = os.path.join(base_output_dir, hospital, modality)
    os.makedirs(output_dir, exist_ok=True)

    roi_list = ['bladder', 'pelvic_cavity', 'pelvic_fat', 'rectum', 'trigone_of_bladder', 'ureter']
    excel_map = {}
    for roi in roi_list:
        path = os.path.join(output_dir, f"{roi}.xlsx")
        if not os.path.exists(path):
            pd.DataFrame(columns=["ImagePath", "MaskPath"]).to_excel(path, index=False)
        excel_map[roi] = path
    return excel_map


def process_case(dirpath, files, modality, hospital, excel_map, start_row=1):
    """Process all ROIs for one case."""
    roi_list = ['bladder', 'pelvic_cavity', 'pelvic_fat', 'rectum', 'trigone_of_bladder', 'ureter']
    image_path = [f for f in files if 'image' in f.lower()][0]
    mask_files = [f for f in files if 'image' not in f.lower()]
    mask_files.sort()

    is_first = True
    is_fact = 0  # Set to 1 if pelvic fat hypertrophy
    for i, roi in enumerate(roi_list):
        mask_path = mask_files[i] if i < len(mask_files) else None
        if mask_path and os.path.exists(mask_path):
            extract_radiomics_features(mask_path, image_path, is_first, modality, hospital, excel_map[roi], is_fact, start_row)
            start_row += 2 if is_first else 1
            is_first = False
    print(f"[Done] {hospital} | {modality} | Case: {os.path.basename(dirpath)}")


def process_hospital_modality(hospital, modality, root_base_dir):
    """Process all cases for one hospital and modality."""
    root_dir = os.path.join(root_base_dir, hospital, modality)
    if not os.path.exists(root_dir):
        print(f"[Skip] {root_dir} not found")
        return

    dir_files = find_3dSlice_files(root_dir)
    excel_map = writer_map(hospital, modality)
    for dirpath, files in dir_files.items():
        process_case(dirpath, files, modality, hospital, excel_map)


# ---------------------------
# Main
# ---------------------------

if __name__ == "__main__":
    root_base_dir = "/mnt/external_drive/PelvicFatData"
    hospitals = ["Peking_First_Hospital", "Jiankong_Hospital", "Emergency_Hospital", "Miyun_Hospital", "Fuxing_Hospital"]
    modalities = ["CT", "MRI"]

    with ProcessPoolExecutor(max_workers=6) as executor:
        for hospital in hospitals:
            for modality in modalities:
                executor.submit(process_hospital_modality, hospital, modality, root_base_dir)

    print("All hospitals and modalities processed successfully.")
